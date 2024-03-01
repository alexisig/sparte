import logging

from django.contrib.postgres.search import TrigramSimilarity
from django.core.management.base import BaseCommand
from django.db.models import Q
from openpyxl import load_workbook

from public_data.models import Scot
from public_data.storages import DataStorage


logger = logging.getLogger("management.commands")


class Command(BaseCommand):
    help = "Use last SCoT bilan to update SIREN."

    def add_arguments(self, parser) -> None:
        parser.add_argument(
            "--all",
            action="store_true",
            help="Refresh for all SCoT, not only those without SIREN",
        )

        parser.add_argument(
            "--similarity",
            type=float,
            help="Refresh for all SCoT, not only those without SIREN",
        )

    def handle(self, *args, **options) -> None:
        logger.info("Start updating SCoT SIREN")

        if options.get("all"):
            Scot.objects.update(siren=None)

        scot_queryset = Scot.objects.filter(Q(siren__isnull=True) | Q(siren__startswith="Unknown"))

        logger.info("SCoTs to process: %d", scot_queryset.count())

        with DataStorage().open("2023-scot.xlsx") as file_stream:
            workbook = load_workbook(file_stream, data_only=True)
        sheet = workbook["Liste des SCoT"]

        siren_mapping = {}
        # data are available from line 5
        for row in sheet.iter_rows(min_row=5, min_col=5, max_col=17):
            name = row[0].value  # col E contains SCoT's name
            siren = row[12].value  # col Q contains SIREN
            if name is not None and siren is not None:
                siren_mapping[name] = siren

        count_found = 0
        for name, siren in siren_mapping.items():
            qs = (
                scot_queryset.annotate(similarity=TrigramSimilarity("name__unaccent", name))
                .filter(similarity__gt=options.get("similarity", 0.8))
                .order_by("-similarity")
            )
            if qs.exists():
                count_found += 1
                scot = qs.first()
                logger.info("Needle: %s ; %s updated with SIREN %s", name, scot.name, siren)
                scot.siren = siren
                scot.save(update_fields=["siren"])

        logger.info("End updating SCoT SIREN. %d updated", count_found)
